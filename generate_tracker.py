import sys
import subprocess

def install(package):
    subprocess.check_call([sys.executable, "-m", "pip", "install", package])

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    install('openpyxl')
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def generate_tracker():
    xlsx_file = r'c:\Users\Vaibhav\DSA-Java\LeetCode_DSA_Tracker.xlsx'

    data = [
        ["No.", "Topic", "What to Learn First", "Problem", "Algorithm Used", "LeetCode Link", "Difficulty", "Solved Alone?", "Date Solved", "Revision 1", "Revision 2", "Revision 3"],
        
        # 0. The Absolute Basics (Brand New)
        [1, "Absolute Basics", "Writing a return statement", "Add Two Integers", "Math", "https://leetcode.com/problems/add-two-integers/", "Easy", "YES", "31-07-2026", "01-08-2026", "07-08-2026", ""],
        [2, "Absolute Basics", "Basic Math Formulas", "Convert the Temperature", "Math", "https://leetcode.com/problems/convert-the-temperature/", "Easy", "", "", "", "", ""],
        [3, "Absolute Basics", "If/Else Statements", "Smallest Even Multiple", "Math", "https://leetcode.com/problems/smallest-even-multiple/", "Easy", "", "", "", "", ""],
        
        # 1. Basic Math & Loops
        [4, "Math & Basics", "Basic Math & Loops", "Subtract the Product and Sum of Digits of an Integer", "-", "https://leetcode.com/problems/subtract-the-product-and-sum-of-digits-of-an-integer/", "Easy", "YES", "28-07-2026", "29-07-2026", "04-08-2026", ""],
        [5, "Math & Basics", "Basic Math & Loops", "Number of Steps to Reduce a Number to Zero", "-", "https://leetcode.com/problems/number-of-steps-to-reduce-a-number-to-zero/", "Easy", "YES", "28-07-2026", "29-07-2026", "04-08-2026", ""],
        [6, "Math & Basics", "Basic Math & Loops", "Power of Two", "Bitwise or Repeated Division", "https://leetcode.com/problems/power-of-two/", "Easy", "YES", "28-07-2026", "29-07-2026", "04-08-2026", ""],
        [7, "Math & Basics", "Basic Math & Loops", "Fibonacci Number", "Recursion or DP", "https://leetcode.com/problems/fibonacci-number/", "Easy", "YES", "22-07-2026", "23-07-2026", "29-07-2026", ""],
        [8, "Math & Basics", "Basic Math & Loops", "Climbing Stairs", "Fibonacci Sequence", "https://leetcode.com/problems/climbing-stairs/", "Easy", "YES", "22-07-2026", "23-07-2026", "29-07-2026", ""],
        [9, "Math & Basics", "Basic Math & Loops", "Palindrome Number", "Digit Extraction", "https://leetcode.com/problems/palindrome-number/", "Easy", "YES", "23-07-2026", "24-07-2026", "30-07-2026", ""],
        [10, "Math & Basics", "Basic Math & Loops", "Reverse Integer", "Digit Extraction", "https://leetcode.com/problems/reverse-integer/", "Medium", "YES", "23-07-2026", "24-07-2026", "30-07-2026", ""],
        [11, "Math & Basics", "Basic Math & Loops", "Count Primes", "Sieve of Eratosthenes", "https://leetcode.com/problems/count-primes/", "Medium", "YES", "23-07-2026", "24-07-2026", "30-07-2026", ""],
        [12, "Math & Basics", "Basic Math & Loops", "Base 7", "Base Conversion", "https://leetcode.com/problems/base-7/", "Easy", "YES", "24-07-2026", "25-07-2026", "31-07-2026", ""],
        [13, "Math & Basics", "Basic Math & Loops", "Add Binary", "-", "https://leetcode.com/problems/add-binary/", "Easy", "YES", "24-07-2026", "25-07-2026", "31-07-2026", ""],
        
        # 1.5 Strings (Beginner Friendly)
        [14, "Strings", "String Manipulation", "Score of a String", "ASCII Values", "https://leetcode.com/problems/score-of-a-string/", "Easy", "", "", "", "", ""],
        [15, "Strings", "String Manipulation", "Defanging an IP Address", "String Replace", "https://leetcode.com/problems/defanging-an-ip-address/", "Easy", "", "", "", "", ""],
        [16, "Strings", "String Manipulation", "Goal Parser Interpretation", "String Replace", "https://leetcode.com/problems/goal-parser-interpretation/", "Easy", "", "", "", "", ""],

        # 2. 1D Arrays (Basic Traversal)
        [17, "Arrays", "Basic Array Traversal", "Build Array from Permutation", "-", "https://leetcode.com/problems/build-array-from-permutation/", "Easy", "", "", "", "", ""],
        [18, "Arrays", "Basic Array Traversal", "Concatenation of Array", "-", "https://leetcode.com/problems/concatenation-of-array/", "Easy", "YES", "30-07-2026", "31-07-2026", "06-08-2026", ""],
        [19, "Arrays", "Basic Array Traversal", "Shuffle the Array", "Array Indexing", "https://leetcode.com/problems/shuffle-the-array/", "Easy", "", "", "", "", ""],
        [20, "Arrays", "Basic Array Traversal", "Kids With the Greatest Number of Candies", "Find Max & Loop", "https://leetcode.com/problems/kids-with-the-greatest-number-of-candies/", "Easy", "", "", "", "", ""],
        [21, "Arrays", "Basic Array Traversal", "Richest Customer Wealth", "2D Array Traversal", "https://leetcode.com/problems/richest-customer-wealth/", "Easy", "", "", "", "", ""],
        [22, "Arrays", "Basic Array Traversal", "Running Sum of 1d Array", "Prefix Sum", "https://leetcode.com/problems/running-sum-of-1d-array/", "Easy", "", "", "", "", ""],
        [23, "Arrays", "Basic Array Traversal", "Find Pivot Index", "Prefix Sum", "https://leetcode.com/problems/find-pivot-index/", "Easy", "", "", "", "", ""],
        [24, "Arrays", "Basic Array Traversal", "Missing Number", "Gauss Formula or XOR", "https://leetcode.com/problems/missing-number/", "Easy", "YES", "22-07-2026", "23-07-2026", "29-07-2026", ""],
        [25, "Arrays", "Basic Array Traversal", "Add to Array-Form of Integer", "-", "https://leetcode.com/problems/add-to-array-form-of-integer/", "Easy", "", "", "", "", ""],
        [26, "Arrays", "Basic Array Traversal", "Best Time to Buy and Sell Stock", "Track Min Value", "https://leetcode.com/problems/best-time-to-buy-and-sell-stock/", "Easy", "YES", "29-07-2026", "30-07-2026", "05-08-2026", ""],
        
        # 3. Array - Two Pointers
        [27, "Arrays - Two Pointers", "Two Pointers Technique", "Reverse String", "Two Pointers (Left/Right)", "https://leetcode.com/problems/reverse-string/", "Easy", "", "", "", "", ""],
        [28, "Arrays - Two Pointers", "Two Pointers Technique", "Remove Element", "Two Pointers (Slow/Fast)", "https://leetcode.com/problems/remove-element/", "Easy", "", "", "", "", ""],
        [29, "Arrays - Two Pointers", "Two Pointers Technique", "Move Zeroes", "Two Pointers (Slow/Fast)", "https://leetcode.com/problems/move-zeroes/", "Easy", "YES", "25-07-2026", "26-07-2026", "01-08-2026", ""],
        [30, "Arrays - Two Pointers", "Two Pointers Technique", "Squares of a Sorted Array", "Two Pointers (Left/Right)", "https://leetcode.com/problems/squares-of-a-sorted-array/", "Easy", "YES", "26-07-2026", "27-07-2026", "02-08-2026", ""],
        
        # 4. Binary Search & Sorting
        [31, "Binary Search & Sort", "Binary Search", "Find Target Indices After Sorting Array", "Sorting", "https://leetcode.com/problems/find-target-indices-after-sorting-array/", "Easy", "", "", "", "", ""],
        [32, "Binary Search & Sort", "Binary Search", "Binary Search", "Binary Search", "https://leetcode.com/problems/binary-search/", "Easy", "", "", "", "", ""],
        [33, "Binary Search & Sort", "Binary Search", "Search Insert Position", "Binary Search", "https://leetcode.com/problems/search-insert-position/", "Easy", "", "", "", "", ""],
        [34, "Binary Search & Sort", "Binary Search", "First Bad Version", "Binary Search", "https://leetcode.com/problems/first-bad-version/", "Easy", "", "", "", "", ""],
        
        # 5. Hashing / HashMaps
        [35, "Hashing", "HashMaps & HashSets", "Jewels and Stones", "HashSet", "https://leetcode.com/problems/jewels-and-stones/", "Easy", "", "", "", "", ""],
        [36, "Hashing", "HashMaps & HashSets", "How Many Numbers Are Smaller Than the Current Number", "Frequency Array", "https://leetcode.com/problems/how-many-numbers-are-smaller-than-the-current-number/", "Easy", "", "", "", "", ""],
        [37, "Hashing", "HashMaps & HashSets", "Contains Duplicate", "HashSet", "https://leetcode.com/problems/contains-duplicate/", "Easy", "", "", "", "", ""],
        [38, "Hashing", "HashMaps & HashSets", "Valid Anagram", "HashMap / Frequency Array", "https://leetcode.com/problems/valid-anagram/", "Easy", "", "", "", "", ""],
        [39, "Hashing", "HashMaps & HashSets", "Two Sum", "HashMap", "https://leetcode.com/problems/two-sum/", "Easy", "YES", "25-07-2026", "26-07-2026", "01-08-2026", ""],
        [40, "Hashing", "HashMaps & HashSets", "Majority Element", "Moore's Voting Algorithm", "https://leetcode.com/problems/majority-element/", "Easy", "YES", "27-07-2026", "28-07-2026", "03-08-2026", ""],
        
        # 6. Basic Stacks
        [41, "Stacks", "Stack Data Structure", "Valid Parentheses", "Stack", "https://leetcode.com/problems/valid-parentheses/", "Easy", "", "", "", "", ""],
        [42, "Stacks", "Stack Data Structure", "Remove Outermost Parentheses", "Stack or Counter", "https://leetcode.com/problems/remove-outermost-parentheses/", "Easy", "", "", "", "", ""],
        [43, "Stacks", "Stack Data Structure", "Remove All Adjacent Duplicates In String", "Stack", "https://leetcode.com/problems/remove-all-adjacent-duplicates-in-string/", "Easy", "", "", "", "", ""],
        [44, "Stacks", "Stack Data Structure", "Min Stack", "Two Stacks", "https://leetcode.com/problems/min-stack/", "Medium", "", "", "", "", ""],
        [45, "Stacks", "Stack Data Structure", "Evaluate Reverse Polish Notation", "Stack", "https://leetcode.com/problems/evaluate-reverse-polish-notation/", "Medium", "", "", "", "", ""],
        [46, "Stacks", "Stack Data Structure", "Next Greater Element I", "Monotonic Stack", "https://leetcode.com/problems/next-greater-element-i/", "Easy", "", "", "", "", ""],
        [47, "Stacks", "Stack Data Structure", "Daily Temperatures", "Monotonic Stack", "https://leetcode.com/problems/daily-temperatures/", "Medium", "", "", "", "", ""]
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "DSA Tracker"

    # Write Data
    for row in data:
        ws.append(row)

    # Styles
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    
    thin_border = Border(left=Side(style='thin', color='BFBFBF'), 
                         right=Side(style='thin', color='BFBFBF'), 
                         top=Side(style='thin', color='BFBFBF'), 
                         bottom=Side(style='thin', color='BFBFBF'))

    # Format header
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    # Format rows
    alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        for cell in row:
            cell.border = thin_border
            # Alternating row colors
            if row_idx % 2 == 0:
                cell.fill = alt_fill
            
            # Formatting difficulty column (now at column 7)
            if cell.column == 7:
                if cell.value == "Easy":
                    cell.font = Font(color="00B050", bold=True)
                elif cell.value == "Medium":
                    cell.font = Font(color="E36C09", bold=True)
                elif cell.value == "Hard":
                    cell.font = Font(color="FF0000", bold=True)
                    
            # LeetCode links (now at column 6)
            if cell.column == 6 and cell.value and isinstance(cell.value, str) and cell.value.startswith("http"):
                cell.hyperlink = cell.value
                cell.font = Font(color="0563C1", underline="single")
            
            # Alignments (No. is at 1)
            if cell.column in [2, 3, 4, 5, 6, 8]:
                cell.alignment = left_align
            else:
                cell.alignment = center_align

    # Column Widths
    widths = {
        'A': 5, 'B': 15, 'C': 20, 'D': 45, 'E': 25, 'F': 45, 'G': 15,
        'H': 15, 'I': 15, 'J': 18, 'K': 18, 'L': 18, 'M': 18
    }
    for col, width in widths.items():
        if col in ws.column_dimensions:
            ws.column_dimensions[col].width = width
        else:
            ws.column_dimensions[col].width = width

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    try:
        wb.save(xlsx_file)
        print("Successfully created formatted Excel file.")
    except PermissionError:
        print("ERROR_OPEN")

if __name__ == "__main__":
    generate_tracker()

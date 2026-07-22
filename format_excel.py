import csv
import subprocess
import sys
import os

def install(package):
    subprocess.check_call([sys.executable, "-m", "pip", "install", package])

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    install('openpyxl')
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

def format_tracker():
    csv_file = r'c:\Users\Vaibhav\DSA-Java\LeetCode_DSA_Tracker.csv'
    xlsx_file = r'c:\Users\Vaibhav\DSA-Java\LeetCode_DSA_Tracker.xlsx'

    wb = Workbook()
    ws = wb.active
    ws.title = "DSA Tracker"

    # Read CSV
    with open(csv_file, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        for row in reader:
            ws.append(row)

    # Styles
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    
    thin_border = Border(left=Side(style='thin', color='BFBFBF'), 
                         right=Side(style='thin', color='BFBFBF'), 
                         top=Side(style='thin', color='BFBFBF'), 
                         bottom=Side(style='thin', color='BFBFBF'))

    # Format header
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    # Format rows
    alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        for cell in row:
            cell.border = thin_border
            # Alternating row colors
            if row_idx % 2 == 0:
                cell.fill = alt_fill
            
            # Formatting difficulty column
            if cell.column == 5:
                if cell.value == "Easy":
                    cell.font = Font(color="00B050", bold=True)
                elif cell.value == "Medium":
                    cell.font = Font(color="E36C09", bold=True)
                elif cell.value == "Hard":
                    cell.font = Font(color="FF0000", bold=True)
                    
            # LeetCode links
            if cell.column == 4 and cell.value and cell.value.startswith("http"):
                cell.hyperlink = cell.value
                cell.font = Font(color="0563C1", underline="single")
            
            # Alignments
            if cell.column in [1, 2, 3, 4, 6]:
                cell.alignment = left_align
            else:
                cell.alignment = center_align

    # Column Widths
    widths = {
        'A': 15, 'B': 20, 'C': 40, 'D': 45, 'E': 15,
        'F': 40, 'G': 15, 'H': 18, 'I': 15, 'J': 18, 'K': 18, 'L': 18
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # Add filters
    ws.auto_filter.ref = ws.dimensions

    # Freeze top row
    ws.freeze_panes = "A2"

    wb.save(xlsx_file)
    print("Successfully created formatted Excel file.")

if __name__ == "__main__":
    format_tracker()
